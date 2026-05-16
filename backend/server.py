from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import csv
import io
from openpyxl import Workbook
import base64
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table as RLTable, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== MODELS ====================

class ClassSection(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    className: str
    sections: List[str]
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClassSectionCreate(BaseModel):
    className: str
    sections: List[str]

class FeeType(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    feeName: str
    amount: float
    applicableClass: Optional[str] = None
    applicableSection: Optional[str] = None
    noticeStartDate: Optional[str] = None
    dueDate: Optional[str] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FeeTypeCreate(BaseModel):
    feeName: str
    amount: float
    applicableClass: Optional[str] = None
    applicableSection: Optional[str] = None
    noticeStartDate: Optional[str] = None
    dueDate: Optional[str] = None

class DatabaseSettings(BaseModel):
    mongoUrl: str
    dbName: str

class Student(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    studentCode: str = ""  # Unique admission/student ID (e.g., ADM001)
    studentName: str
    rollNo: str  # Class-wise roll number (not unique)
    studentClass: str
    section: str
    fatherName: str
    motherName: str
    mobile: str
    address: str
    feeTerm1: float
    feeTerm2: float
    feeTerm3: float
    parentUsername: Optional[str] = None
    parentPassword: Optional[str] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StudentCreate(BaseModel):
    studentCode: str
    studentName: str
    rollNo: str
    studentClass: str
    section: str
    fatherName: str
    motherName: str
    mobile: str
    address: str
    feeTerm1: float
    feeTerm2: float
    feeTerm3: float
    parentUsername: Optional[str] = None
    parentPassword: Optional[str] = None

class StudentUpdate(BaseModel):
    studentCode: Optional[str] = None
    studentName: Optional[str] = None
    rollNo: Optional[str] = None
    studentClass: Optional[str] = None
    section: Optional[str] = None
    fatherName: Optional[str] = None
    motherName: Optional[str] = None
    mobile: Optional[str] = None
    address: Optional[str] = None
    feeTerm1: Optional[float] = None
    feeTerm2: Optional[float] = None
    feeTerm3: Optional[float] = None
    parentUsername: Optional[str] = None
    parentPassword: Optional[str] = None

class AttendanceRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    studentId: str
    rollNo: str
    studentName: str
    studentClass: str
    section: str
    date: str
    status: str
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class AttendanceSubmit(BaseModel):
    studentClass: str
    section: str
    date: str
    records: List[Dict]

class FeePayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    studentId: str
    studentCode: str
    rollNo: str
    studentName: str
    termNumber: Optional[int] = None
    feeTypeId: Optional[str] = None
    feeName: Optional[str] = None
    amount: float
    paymentMode: str
    upiScreenshot: Optional[str] = None
    receiptNumber: str
    collectedBy: Optional[str] = None
    paymentDate: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FeePaymentCreate(BaseModel):
    studentId: str
    studentCode: str
    rollNo: str
    studentName: str
    termNumber: Optional[int] = None
    feeTypeId: Optional[str] = None
    feeName: Optional[str] = None
    amount: float
    paymentMode: str
    upiScreenshot: Optional[str] = None
    collectedBy: Optional[str] = None

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    expenseName: str
    amount: float
    date: str
    billUrl: str
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ExpenseCreate(BaseModel):
    expenseName: str
    amount: float
    date: str
    billUrl: str

class WhatsAppSettings(BaseModel):
    phoneNumberId: str
    accessToken: str

class SchoolSettings(BaseModel):
    schoolName: str
    schoolAddress: str
    logoUrl: Optional[str] = None

class PromoteRequest(BaseModel):
    fromClass: str
    toClass: str

class InventoryItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    itemName: str
    quantity: int
    category: str
    purchaseDate: str
    amount: float
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InventoryItemCreate(BaseModel):
    itemName: str
    quantity: int
    category: str
    purchaseDate: str
    amount: float

class InventoryIssue(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    itemId: str
    itemName: str
    studentId: str
    studentCode: str
    rollNo: str
    studentName: str
    quantity: int
    date: str
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class InventoryIssueCreate(BaseModel):
    itemId: str
    studentCode: str
    quantity: int
    date: str

class Event(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: str
    date: str
    sendNotification: Optional[bool] = False
    attachmentUrl: Optional[str] = None
    attachmentName: Optional[str] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EventCreate(BaseModel):
    title: str
    description: str
    date: str
    sendNotification: Optional[bool] = False
    attachmentUrl: Optional[str] = None
    attachmentName: Optional[str] = None

class Homework(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    studentClass: str
    section: str
    subject: str
    title: str
    description: str
    dueDate: str
    assignedBy: str
    attachmentUrl: Optional[str] = None
    attachmentName: Optional[str] = None
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class HomeworkCreate(BaseModel):
    studentClass: str
    section: str
    subject: str
    title: str
    description: str
    dueDate: str
    assignedBy: str
    attachmentUrl: Optional[str] = None
    attachmentName: Optional[str] = None

class Staff(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    role: str  # teacher, office_staff
    mobile: str
    subject: Optional[str] = None
    joiningDate: str
    username: str
    password: str
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StaffCreate(BaseModel):
    name: str
    role: str
    mobile: str
    subject: Optional[str] = None
    joiningDate: str
    username: str
    password: str

class StaffUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    mobile: Optional[str] = None
    subject: Optional[str] = None
    joiningDate: Optional[str] = None
    password: Optional[str] = None

class LoginRequest(BaseModel):
    username: str
    password: str

class Concession(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    studentId: str
    studentCode: str
    studentName: str
    termNumber: Optional[int] = None
    feeTypeId: Optional[str] = None
    feeName: Optional[str] = None
    concessionAmount: float
    letterUrl: Optional[str] = None
    requestedBy: str
    status: str = "pending"  # pending, approved, rejected
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ConcessionCreate(BaseModel):
    studentCode: str
    termNumber: Optional[int] = None
    feeTypeId: Optional[str] = None
    feeName: Optional[str] = None
    concessionAmount: float
    letterUrl: Optional[str] = None
    requestedBy: str

class BulkConcessionCreate(BaseModel):
    studentCodes: List[str]
    termNumber: Optional[int] = None
    feeTypeId: Optional[str] = None
    feeName: Optional[str] = None
    concessionAmount: float
    letterUrl: Optional[str] = None
    requestedBy: str

class LeaveRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    studentId: str
    studentCode: str
    studentName: str
    fromDate: str
    toDate: str
    reason: str
    attachmentUrl: Optional[str] = None
    status: str = "pending"
    createdAt: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeaveRequestCreate(BaseModel):
    studentId: str
    studentCode: str
    studentName: str
    fromDate: str
    toDate: str
    reason: str
    attachmentUrl: Optional[str] = None

class MarkEntry(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    studentId: str
    studentCode: str
    studentName: str
    studentClass: str
    section: str
    examName: str
    subject: str
    marks: float
    maxMarks: float = 100
    recordedBy: str = ""
    recordedOn: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MarkRow(BaseModel):
    studentCode: str
    studentName: Optional[str] = ""
    examName: str
    subject: str
    marks: float
    maxMarks: Optional[float] = 100

class MarksBulkCreate(BaseModel):
    studentClass: str
    section: str
    examName: str
    subject: str
    maxMarks: Optional[float] = 100
    recordedBy: Optional[str] = ""
    rows: List[MarkRow]


# ==================== WHATSAPP SERVICE ====================

BASE_WA_URL = "https://crm.abhiit.com/api/meta/v19.0"

async def get_wa_settings():
    settings = await db.settings.find_one({"type": "whatsapp"}, {"_id": 0})
    if not settings or not settings.get('phoneNumberId') or not settings.get('accessToken'):
        return None
    return settings

async def send_wa_template(mobile, template_name, components, settings=None):
    """Send WhatsApp template message"""
    try:
        if not settings:
            settings = await get_wa_settings()
        if not settings:
            return {"success": False, "message": "WhatsApp not configured"}
        url = f"{BASE_WA_URL}/{settings['phoneNumberId']}/messages"
        headers = {"Authorization": f"Bearer {settings['accessToken']}", "Content-Type": "application/json"}
        payload = {
            "to": mobile,
            "recipient_type": "individual",
            "type": "template",
            "template": {
                "language": {"policy": "deterministic", "code": "en"},
                "name": template_name,
                "components": components
            }
        }
        async with httpx.AsyncClient() as http_client:
            response = await http_client.post(url, headers=headers, json=payload, timeout=30.0)
            logger.info(f"WhatsApp template '{template_name}' sent to {mobile}: {response.status_code}")
            return {"success": True, "data": response.json()}
    except Exception as e:
        logger.error(f"WhatsApp send failed: {str(e)}")
        return {"success": False, "message": str(e)}

async def send_fee_paid_message(mobile, invoice_url, amount, fee_name, student_name, settings=None):
    """Send fee paid success with invoice document"""
    components = [
        {"type": "header", "parameters": [{"type": "document", "document": {"link": invoice_url}}]},
        {"type": "body", "parameters": [
            {"type": "text", "text": str(amount)},
            {"type": "text", "text": fee_name},
            {"type": "text", "text": student_name}
        ]}
    ]
    return await send_wa_template(mobile, "fee_paid_bill", components, settings)

async def send_absent_message(mobile, student_name, class_name, date_str, settings=None):
    """Send absent notification"""
    components = [
        {"type": "body", "parameters": [
            {"type": "text", "text": student_name},
            {"type": "text", "text": class_name},
            {"type": "text", "text": date_str}
        ]}
    ]
    return await send_wa_template(mobile, "absent_hifg", components, settings)

async def send_event_message(mobile, event_name, event_date, settings=None):
    """Send event notification"""
    components = [
        {"type": "body", "parameters": [
            {"type": "text", "text": event_name},
            {"type": "text", "text": event_date}
        ]}
    ]
    return await send_wa_template(mobile, "holi", components, settings)

# Backward-compat wrapper
async def send_whatsapp_message(mobile, message, settings=None):
    """Fallback text message (kept for fee reminders etc)"""
    try:
        if not settings:
            settings = await get_wa_settings()
        if not settings:
            return {"success": False, "message": "WhatsApp not configured"}
        url = f"{BASE_WA_URL}/{settings['phoneNumberId']}/messages"
        headers = {"Authorization": f"Bearer {settings['accessToken']}", "Content-Type": "application/json"}
        payload = {"to": mobile, "recipient_type": "individual", "type": "text", "text": {"body": message}}
        async with httpx.AsyncClient() as http_client:
            response = await http_client.post(url, headers=headers, json=payload, timeout=30.0)
            return {"success": True, "data": response.json()}
    except Exception as e:
        logger.error(f"WhatsApp send failed: {str(e)}")
        return {"success": False, "message": str(e)}

# ==================== PDF INVOICE GENERATION ====================

def generate_invoice_pdf(payment_data, student_data, school_settings=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=10*mm, bottomMargin=8*mm, leftMargin=12*mm, rightMargin=12*mm)
    styles = getSampleStyleSheet()
    elements = []

    school_name = (school_settings or {}).get('schoolName', 'SchoolPro')
    school_address = (school_settings or {}).get('schoolAddress', '')

    purple_dark = colors.HexColor('#6B21A8')
    purple_light = colors.HexColor('#9333EA')
    gray_text = colors.HexColor('#6B7280')
    light_border = colors.HexColor('#E5E7EB')

    receipt_id = payment_data.get('receiptNumber', '')
    pay_date = payment_data.get('paymentDate', '')
    if isinstance(pay_date, str) and len(pay_date) >= 10:
        try:
            parts = pay_date[:10].split('-')
            pay_date = f"{parts[2]}-{parts[1]}-{parts[0]}"
        except Exception:
            pay_date = pay_date[:10]
    else:
        pay_date = datetime.now().strftime('%d-%m-%Y')

    student_code = student_data.get('studentCode', student_data.get('rollNo', ''))
    student_name = student_data.get('studentName', '')
    fee_label = f"Term {payment_data.get('termNumber')}" if payment_data.get('termNumber') else (payment_data.get('feeName') or 'Custom Fee')
    amount = payment_data.get('amount', 0)
    collected_by = payment_data.get('collectedBy', 'Admin')
    payment_mode = payment_data.get('paymentMode', '').upper()

    def build_receipt_copy(copy_type):
        """Build one receipt copy (Student/College)"""
        copy_elements = []

        # School Name
        name_style = ParagraphStyle('SN', parent=styles['Title'], fontSize=18, textColor=purple_dark, alignment=1, spaceAfter=1, leading=20)
        copy_elements.append(Paragraph(school_name, name_style))
        if school_address:
            addr_style = ParagraphStyle('SA', parent=styles['Normal'], fontSize=8, textColor=gray_text, alignment=1, spaceAfter=2)
            copy_elements.append(Paragraph(school_address, addr_style))
        copy_elements.append(Spacer(1, 2*mm))

        # Header bar: Receipt ID | COPY TYPE FEE RECEIPT | Date
        hdr_label = "STUDENT FEE RECEIPT" if copy_type == "student" else "COLLEGE FEE RECEIPT"
        hdr = [[f"Receipt ID: #{receipt_id}", hdr_label, f"Date: {pay_date}"]]
        ht = RLTable(hdr, colWidths=[155, 220, 115])
        ht.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), purple_dark),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'), ('ALIGN', (1, 0), (1, 0), 'CENTER'), ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, 0), 8), ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('LEFTPADDING', (0, 0), (-1, 0), 8), ('RIGHTPADDING', (0, 0), (-1, 0), 8),
        ]))
        copy_elements.append(ht)
        copy_elements.append(Spacer(1, 1*mm))

        # Table header
        th = [["STUDENT NAME & ID", "FEE TYPE", "AMOUNT PAID"]]
        tht = RLTable(th, colWidths=[200, 170, 120])
        tht.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), purple_light),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('TOPPADDING', (0, 0), (-1, 0), 7), ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
            ('LEFTPADDING', (0, 0), (-1, 0), 8), ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
        ]))
        copy_elements.append(tht)

        # Table data
        td = [[f"{student_code} - {student_name}", fee_label, f"Rs. {amount:,.2f}"]]
        tdt = RLTable(td, colWidths=[200, 170, 120])
        tdt.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'), ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 8), ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 8), ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('BOX', (0, 0), (-1, -1), 0.5, light_border),
        ]))
        copy_elements.append(tdt)
        copy_elements.append(Spacer(1, 2*mm))

        # Additional details row
        detail_data = [
            [f"Class: {student_data.get('studentClass', '')} - {student_data.get('section', '')}", f"Father: {student_data.get('fatherName', '')}", f"Mode: {payment_mode}"],
        ]
        ddt = RLTable(detail_data, colWidths=[170, 180, 140])
        ddt.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 8), ('TEXTCOLOR', (0, 0), (-1, -1), gray_text),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ]))
        copy_elements.append(ddt)
        copy_elements.append(Spacer(1, 2*mm))

        # Note
        n = ParagraphStyle('N', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#9CA3AF'), alignment=0)
        copy_elements.append(Paragraph("Note: Some Times Fee Payments Take Time To Update In Our System", n))
        copy_elements.append(Spacer(1, 2*mm))

        # Computer generated + Collected by
        g = ParagraphStyle('G', parent=styles['Normal'], fontSize=8, textColor=gray_text, alignment=1)
        copy_elements.append(Paragraph("This is a computer-generated invoice and does not require a physical signature.", g))
        copy_elements.append(Spacer(1, 1*mm))
        copy_elements.append(Paragraph(f"Processed By: {collected_by}", g))

        return copy_elements

    # Build Student Copy
    elements.extend(build_receipt_copy("student"))

    # Separator line
    elements.append(Spacer(1, 4*mm))
    sep_data = [["- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -"]]
    sep = RLTable(sep_data, colWidths=[490])
    sep.setStyle(TableStyle([('ALIGN', (0, 0), (0, 0), 'CENTER'), ('TEXTCOLOR', (0, 0), (0, 0), colors.HexColor('#D1D5DB')), ('FONTSIZE', (0, 0), (0, 0), 7)]))
    elements.append(sep)
    elements.append(Spacer(1, 4*mm))

    # Build College Copy
    elements.extend(build_receipt_copy("college"))

    # Footer
    elements.append(Spacer(1, 3*mm))
    f = ParagraphStyle('F', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#D1D5DB'), alignment=1)
    elements.append(Paragraph("Software Designed & Developed By SchoolPro", f))

    doc.build(elements)
    buf.seek(0)
    return buf

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/login")
async def login(data: LoginRequest):
    # Check super admin
    if data.username == "admin" and data.password == "12345678":
        return {"success": True, "user": {"name": "Super Admin", "username": "admin"}, "role": "super_admin"}
    # Check staff (teacher, office_staff, admin_role)
    staff = await db.staff.find_one({"username": data.username, "password": data.password}, {"_id": 0})
    if staff:
        return {"success": True, "user": {k: v for k, v in staff.items() if k != 'password'}, "role": staff['role']}
    raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.post("/auth/staff-login")
async def staff_login(data: LoginRequest):
    staff = await db.staff.find_one({"username": data.username, "password": data.password}, {"_id": 0})
    if not staff:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    return {"success": True, "user": {k: v for k, v in staff.items() if k != 'password'}, "role": staff['role']}

@api_router.post("/auth/parent-login")
async def parent_login(data: LoginRequest):
    student = await db.students.find_one({"parentUsername": data.username, "parentPassword": data.password}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    return {"success": True, "student": {k: v for k, v in student.items() if k != 'parentPassword'}, "role": "parent"}

# ==================== CLASS & SECTION ROUTES ====================

@api_router.post("/classes", response_model=ClassSection)
async def create_class_section(data: ClassSectionCreate):
    existing = await db.classes.find_one({"className": data.className}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Class already exists")
    obj = ClassSection(**data.model_dump())
    doc = obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.classes.insert_one(doc)
    return obj

@api_router.get("/classes")
async def get_classes():
    return await db.classes.find({}, {"_id": 0}).to_list(100)

@api_router.put("/classes/{class_id}")
async def update_class_section(class_id: str, data: ClassSectionCreate):
    result = await db.classes.update_one({"id": class_id}, {"$set": {"className": data.className, "sections": data.sections}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Class not found")
    return await db.classes.find_one({"id": class_id}, {"_id": 0})

@api_router.delete("/classes/{class_id}")
async def delete_class_section(class_id: str):
    result = await db.classes.delete_one({"id": class_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Class not found")
    return {"message": "Class deleted"}

# ==================== FEE TYPES ROUTES ====================

@api_router.post("/fee-types", response_model=FeeType)
async def create_fee_type(data: FeeTypeCreate):
    obj = FeeType(**data.model_dump())
    doc = obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.fee_types.insert_one(doc)
    return obj

@api_router.get("/fee-types")
async def get_fee_types(applicableClass: Optional[str] = None):
    query = {}
    if applicableClass:
        query['$or'] = [{'applicableClass': applicableClass}, {'applicableClass': None}, {'applicableClass': ''}]
    return await db.fee_types.find(query, {"_id": 0}).to_list(500)

@api_router.put("/fee-types/{fee_type_id}")
async def update_fee_type(fee_type_id: str, data: FeeTypeCreate):
    result = await db.fee_types.update_one({"id": fee_type_id}, {"$set": data.model_dump()})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Fee type not found")
    return await db.fee_types.find_one({"id": fee_type_id}, {"_id": 0})

@api_router.delete("/fee-types/{fee_type_id}")
async def delete_fee_type(fee_type_id: str):
    result = await db.fee_types.delete_one({"id": fee_type_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Fee type not found")
    return {"message": "Fee type deleted"}

# ==================== DATABASE SETTINGS ====================

@api_router.get("/settings/database")
async def get_database_settings():
    settings = await db.settings.find_one({"type": "database"}, {"_id": 0})
    if not settings:
        return {"mongoUrl": os.environ.get('MONGO_URL', ''), "dbName": os.environ.get('DB_NAME', '')}
    return settings

@api_router.put("/settings/database")
async def update_database_settings(data: DatabaseSettings):
    global client, db
    # Add authSource=admin if not already in URL for authenticated connections
    mongo_url = data.mongoUrl
    if '@' in mongo_url and 'authSource' not in mongo_url:
        if '?' not in mongo_url:
            if not mongo_url.endswith('/'):
                mongo_url = f"{mongo_url}/"
            mongo_url = f"{mongo_url}?authSource=admin"
        else:
            mongo_url = f"{mongo_url}&authSource=admin"
    try:
        test_client = AsyncIOMotorClient(mongo_url, serverSelectionTimeoutMS=5000)
        await test_client[data.dbName].command('ping')
        test_client.close()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect: {str(e)}")
    # Switch connection
    client.close()
    client = AsyncIOMotorClient(mongo_url)
    db = client[data.dbName]
    # Save settings in the NEW database
    await db.settings.update_one({"type": "database"}, {"$set": {"mongoUrl": data.mongoUrl, "dbName": data.dbName}}, upsert=True)
    return {"message": "Database connected successfully"}

# ==================== STUDENT ROUTES ====================

@api_router.post("/students", response_model=Student)
async def create_student(student: StudentCreate):
    existing = await db.students.find_one({"studentCode": student.studentCode}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Student ID already exists")
    student_obj = Student(**student.model_dump())
    doc = student_obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.students.insert_one(doc)
    return student_obj

@api_router.post("/students/bulk")
async def bulk_upload_students(file: UploadFile = File(...)):
    try:
        content = await file.read()
        decoded = content.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(decoded))
        added, errors = 0, []
        for row in csv_reader:
            try:
                student_data = StudentCreate(
                    studentCode=row['Student ID'].strip(),
                    studentName=row['Student Name'].strip(), rollNo=row['Roll No'].strip(),
                    studentClass=row['Class'].strip(), section=row['Section'].strip(),
                    fatherName=row['Father Name'].strip(), motherName=row['Mother Name'].strip(),
                    mobile=row['Mobile Number'].strip(), address=row['Address'].strip(),
                    feeTerm1=float(row['Fee Term1']), feeTerm2=float(row['Fee Term2']), feeTerm3=float(row['Fee Term3']),
                    parentUsername=row.get('Parent Username', '').strip() or None,
                    parentPassword=row.get('Parent Password', '').strip() or None,
                )
                existing = await db.students.find_one({"studentCode": student_data.studentCode}, {"_id": 0})
                if existing:
                    errors.append(f"Student ID {student_data.studentCode} exists")
                    continue
                student_obj = Student(**student_data.model_dump())
                doc = student_obj.model_dump()
                doc['createdAt'] = doc['createdAt'].isoformat()
                await db.students.insert_one(doc)
                added += 1
            except Exception as e:
                errors.append(f"Row error: {str(e)}")
        return {"added": added, "errors": errors}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@api_router.get("/students/sample-csv")
async def download_sample_csv():
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Student ID', 'Student Name', 'Roll No', 'Class', 'Section', 'Father Name', 'Mother Name', 'Mobile Number', 'Address', 'Fee Term1', 'Fee Term2', 'Fee Term3', 'Parent Username', 'Parent Password'])
    writer.writerow(['ADM001', 'John Doe', '1', '1', 'A', 'Robert Doe', 'Jane Doe', '9876543210', '123 Main St', '5000', '5000', '5000', 'parent101', 'pass101'])
    writer.writerow(['ADM002', 'Alice Smith', '2', '1', 'A', 'Michael Smith', 'Sarah Smith', '9876543211', '456 Oak Ave', '5000', '5000', '5000', 'parent102', 'pass102'])
    output.seek(0)
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=sample_students.csv"})

@api_router.get("/students")
async def get_students(studentClass: Optional[str] = None, section: Optional[str] = None, search: Optional[str] = None, page: int = 1, limit: int = 50):
    query = {}
    if studentClass: query['studentClass'] = studentClass
    if section: query['section'] = section
    if search: query['$or'] = [{'studentName': {'$regex': search, '$options': 'i'}}, {'rollNo': {'$regex': search, '$options': 'i'}}, {'studentCode': {'$regex': search, '$options': 'i'}}]
    total = await db.students.count_documents(query)
    skip = (page - 1) * limit
    students = await db.students.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    for s in students:
        if isinstance(s.get('createdAt'), str): s['createdAt'] = datetime.fromisoformat(s['createdAt'])
        if 'studentCode' not in s: s['studentCode'] = s.get('rollNo', '')
    return {"students": students, "total": total, "page": page, "limit": limit, "totalPages": max(1, -(-total // limit))}

@api_router.put("/students/{student_id}", response_model=Student)
async def update_student(student_id: str, update_data: StudentUpdate):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student: raise HTTPException(status_code=404, detail="Student not found")
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    if update_dict: await db.students.update_one({"id": student_id}, {"$set": update_dict})
    updated = await db.students.find_one({"id": student_id}, {"_id": 0})
    if isinstance(updated.get('createdAt'), str): updated['createdAt'] = datetime.fromisoformat(updated['createdAt'])
    return Student(**updated)

@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str):
    result = await db.students.delete_one({"id": student_id})
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="Student not found")
    return {"message": "Student deleted"}

@api_router.post("/students/promote-preview")
async def promote_students_preview(request: PromoteRequest):
    """Compute new fee structure for all students in fromClass without committing."""
    students = await db.students.find({"studentClass": request.fromClass}, {"_id": 0}).to_list(10000)
    if not students:
        raise HTTPException(status_code=404, detail="No students found")

    preview = []
    for student in students:
        active_payments = await db.fee_payments.find(
            {"studentId": student['id'], "status": {"$nin": ["reverted", "archived"]}}, {"_id": 0}
        ).to_list(1000)
        total_paid = sum(p.get('amount', 0) for p in active_payments)
        old_custom_fees = await db.fee_types.find({
            "$or": [
                {"applicableClass": request.fromClass, "applicableSection": student.get('section', '')},
                {"applicableClass": request.fromClass, "applicableSection": {"$in": [None, ""]}},
                {"applicableClass": {"$in": [None, ""]}, "applicableSection": {"$in": [None, ""]}},
            ]
        }, {"_id": 0}).to_list(500)
        total_custom = sum(cf.get('amount', 0) for cf in old_custom_fees)
        old_t1 = student.get('feeTerm1', 0)
        old_t2 = student.get('feeTerm2', 0)
        old_t3 = student.get('feeTerm3', 0)
        total_expected = old_t1 + old_t2 + old_t3 + total_custom
        total_due = max(0, total_expected - total_paid)
        new_t1 = old_t1 + total_due
        new_t2 = old_t2
        new_t3 = old_t3 + 5000
        preview.append({
            "studentId": student['id'],
            "studentCode": student.get('studentCode', ''),
            "studentName": student.get('studentName', ''),
            "rollNo": student.get('rollNo', ''),
            "section": student.get('section', ''),
            "totalPaid": total_paid,
            "totalExpected": total_expected,
            "totalDue": total_due,
            "oldFees": {"term1": old_t1, "term2": old_t2, "term3": old_t3, "customFeesTotal": total_custom},
            "newFees": {"term1": new_t1, "term2": new_t2, "term3": new_t3},
        })
    return {"fromClass": request.fromClass, "toClass": request.toClass, "studentCount": len(preview), "preview": preview}

@api_router.post("/students/promote")
async def promote_students(request: PromoteRequest):
    students = await db.students.find({"studentClass": request.fromClass}, {"_id": 0}).to_list(10000)
    if not students:
        raise HTTPException(status_code=404, detail="No students found")
    
    promoted_count = 0
    for student in students:
        await _promote_one_student(student, request.toClass)
        promoted_count += 1
    
    return {"message": f"Promoted {promoted_count} students from {request.fromClass} to {request.toClass}. Previous year due added to Term 1, Term 3 increased by Rs.5000."}


class SingleStudentPromote(BaseModel):
    toClass: str

async def _calc_promotion(student: Dict):
    """Shared calc logic. Returns dict with totals + new fees."""
    from_class = student.get('studentClass', '')
    active_payments = await db.fee_payments.find(
        {"studentId": student['id'], "status": {"$nin": ["reverted", "archived"]}}, {"_id": 0}
    ).to_list(1000)
    total_paid = sum(p.get('amount', 0) for p in active_payments)
    old_custom_fees = await db.fee_types.find({
        "$or": [
            {"applicableClass": from_class, "applicableSection": student.get('section', '')},
            {"applicableClass": from_class, "applicableSection": {"$in": [None, ""]}},
            {"applicableClass": {"$in": [None, ""]}, "applicableSection": {"$in": [None, ""]}},
        ]
    }, {"_id": 0}).to_list(500)
    total_custom = sum(cf.get('amount', 0) for cf in old_custom_fees)
    old_t1 = student.get('feeTerm1', 0)
    old_t2 = student.get('feeTerm2', 0)
    old_t3 = student.get('feeTerm3', 0)
    total_expected = old_t1 + old_t2 + old_t3 + total_custom
    total_due = max(0, total_expected - total_paid)
    return {
        "fromClass": from_class,
        "totalPaid": total_paid, "totalExpected": total_expected, "totalDue": total_due,
        "oldFees": {"term1": old_t1, "term2": old_t2, "term3": old_t3, "customFeesTotal": total_custom},
        "newFees": {"term1": old_t1 + total_due, "term2": old_t2, "term3": old_t3 + 5000},
    }

async def _promote_one_student(student: Dict, to_class: str):
    """Promotes a single student using current fee carryover rules. Appends to promotionHistory."""
    calc = await _calc_promotion(student)
    history_entry = {
        "fromClass": calc['fromClass'],
        "toClass": to_class,
        "totalDue": calc['totalDue'],
        "totalPaid": calc['totalPaid'],
        "oldFees": calc['oldFees'],
        "newFees": calc['newFees'],
        "promotedOn": datetime.now(timezone.utc).isoformat(),
    }
    # Build the update
    set_doc = {
        "studentClass": to_class,
        "feeTerm1": calc['newFees']['term1'],
        "feeTerm2": calc['newFees']['term2'],
        "feeTerm3": calc['newFees']['term3'],
        "previousYearDues": {
            "amount": calc['totalDue'],
            "fromClass": calc['fromClass'],
            "promotedOn": history_entry['promotedOn']
        },
        "academicYear": str(datetime.now().year),
    }
    await db.students.update_one(
        {"id": student['id']},
        {"$set": set_doc, "$push": {"promotionHistory": history_entry}}
    )
    # Archive existing payments
    await db.fee_payments.update_many(
        {"studentId": student['id'], "status": {"$nin": ["reverted", "archived"]}},
        {"$set": {"status": "archived"}}
    )

@api_router.post("/students/{student_id}/promote-preview")
async def promote_single_preview(student_id: str, data: SingleStudentPromote):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student: raise HTTPException(status_code=404, detail="Student not found")
    calc = await _calc_promotion(student)
    return {
        "studentId": student['id'],
        "studentCode": student.get('studentCode', ''),
        "studentName": student.get('studentName', ''),
        "rollNo": student.get('rollNo', ''),
        "section": student.get('section', ''),
        "fromClass": calc['fromClass'],
        "toClass": data.toClass,
        "totalPaid": calc['totalPaid'],
        "totalExpected": calc['totalExpected'],
        "totalDue": calc['totalDue'],
        "oldFees": calc['oldFees'],
        "newFees": calc['newFees'],
    }

@api_router.post("/students/{student_id}/promote")
async def promote_single_student(student_id: str, data: SingleStudentPromote):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student: raise HTTPException(status_code=404, detail="Student not found")
    await _promote_one_student(student, data.toClass)
    return {"message": f"{student.get('studentName')} promoted to class {data.toClass}."}

class BulkDeleteRequest(BaseModel):
    studentIds: List[str]

@api_router.post("/students/bulk-delete")
async def bulk_delete_students(data: BulkDeleteRequest):
    if not data.studentIds:
        raise HTTPException(status_code=400, detail="No students selected")
    result = await db.students.delete_many({"id": {"$in": data.studentIds}})
    return {"message": f"Deleted {result.deleted_count} students"}

# ==================== ATTENDANCE ROUTES ====================

@api_router.post("/attendance")
async def mark_attendance(data: AttendanceSubmit):
    await db.attendance.delete_many({"studentClass": data.studentClass, "section": data.section, "date": data.date})
    records = []
    for record in data.records:
        att = AttendanceRecord(studentId=record['studentId'], rollNo=record['rollNo'], studentName=record['studentName'],
                               studentClass=data.studentClass, section=data.section, date=data.date, status=record['status'])
        doc = att.model_dump()
        doc['createdAt'] = doc['createdAt'].isoformat()
        records.append(doc)
    if records: await db.attendance.insert_many(records)
    return {"message": f"Attendance marked for {len(records)} students"}

@api_router.get("/attendance")
async def get_attendance(studentClass: Optional[str] = None, section: Optional[str] = None, startDate: Optional[str] = None, endDate: Optional[str] = None, date: Optional[str] = None, studentId: Optional[str] = None):
    query = {}
    if studentId: query['studentId'] = studentId
    if studentClass: query['studentClass'] = studentClass
    if section: query['section'] = section
    if date: query['date'] = date
    elif startDate and endDate: query['date'] = {'$gte': startDate, '$lte': endDate}
    elif startDate: query['date'] = startDate
    return await db.attendance.find(query, {"_id": 0}).to_list(10000)

@api_router.get("/attendance/export")
async def export_attendance(studentClass: str, section: str, startDate: str, endDate: str, format: str = 'csv'):
    records = await db.attendance.find({"studentClass": studentClass, "section": section, "date": {'$gte': startDate, '$lte': endDate}}, {"_id": 0}).to_list(10000)
    if format == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(['Roll No', 'Student Name', 'Date', 'Status'])
        for r in records: ws.append([r['rollNo'], r['studentName'], r['date'], r['status']])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=attendance.xlsx"})
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Roll No', 'Student Name', 'Date', 'Status'])
    for r in records: writer.writerow([r['rollNo'], r['studentName'], r['date'], r['status']])
    output.seek(0)
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=attendance.csv"})

@api_router.post("/attendance/send-alerts")
async def send_attendance_alerts(data: Dict):
    absent_records = data.get('absentRecords', [])
    settings_doc = await get_wa_settings()
    sent_count = 0
    for record in absent_records:
        class_name = f"{record.get('studentClass', '')}-{record.get('section', '')}"
        result = await send_absent_message(record.get('mobile', ''), record['studentName'], class_name, record['date'], settings_doc)
        if result.get('success'): sent_count += 1
    return {"message": f"Alerts sent to {sent_count} parents"}

@api_router.get("/fees/status")
async def get_fee_status(studentClass: str, section: str):
    """Get fee status for all students in a class/section"""
    students = await db.students.find({"studentClass": studentClass, "section": section}, {"_id": 0}).to_list(1000)
    # Get all custom fee types applicable
    custom_fees = await db.fee_types.find({
        "$or": [
            {"applicableClass": studentClass, "applicableSection": section},
            {"applicableClass": studentClass, "applicableSection": {"$in": [None, ""]}},
            {"applicableClass": {"$in": [None, ""]}, "applicableSection": {"$in": [None, ""]}},
        ]
    }, {"_id": 0}).to_list(500)

    result = []
    for student in students:
        payments = await db.fee_payments.find({"studentId": student['id']}, {"_id": 0}).to_list(100)
        paid_terms = {}
        paid_custom = {}
        for p in payments:
            if p.get('status') in ('reverted', 'archived'): continue
            if p.get('termNumber'):
                k = f"term{p['termNumber']}"
                paid_terms[k] = paid_terms.get(k, 0) + p['amount']
            if p.get('feeTypeId'):
                paid_custom[p['feeTypeId']] = paid_custom.get(p['feeTypeId'], 0) + p['amount']

        total_expected = student.get('feeTerm1', 0) + student.get('feeTerm2', 0) + student.get('feeTerm3', 0)
        total_expected += sum(cf['amount'] for cf in custom_fees)
        total_paid = sum(paid_terms.values()) + sum(paid_custom.values())

        row = {
            "rollNo": student['rollNo'],
            "studentName": student['studentName'],
            "mobile": student.get('mobile', ''),
            "term1Total": student.get('feeTerm1', 0),
            "term1Paid": paid_terms.get('term1', 0),
            "term2Total": student.get('feeTerm2', 0),
            "term2Paid": paid_terms.get('term2', 0),
            "term3Total": student.get('feeTerm3', 0),
            "term3Paid": paid_terms.get('term3', 0),
            "customFees": [{
                "feeName": cf['feeName'],
                "total": cf['amount'],
                "paid": paid_custom.get(cf['id'], 0)
            } for cf in custom_fees],
            "totalExpected": total_expected,
            "totalPaid": total_paid,
            "totalPending": total_expected - total_paid,
        }
        result.append(row)
    return {"students": result, "customFeeNames": [cf['feeName'] for cf in custom_fees]}

@api_router.get("/fees/status/export")
async def export_fee_status(studentClass: str, section: str, format: str = 'csv'):
    data = await get_fee_status(studentClass, section)
    students = data['students']
    custom_names = data['customFeeNames']

    headers = ['Roll No', 'Name', 'Term1 Total', 'Term1 Paid', 'Term2 Total', 'Term2 Paid', 'Term3 Total', 'Term3 Paid']
    for cn in custom_names:
        headers.extend([f'{cn} Total', f'{cn} Paid'])
    headers.extend(['Total Expected', 'Total Paid', 'Total Pending'])

    rows = []
    for s in students:
        row = [s['rollNo'], s['studentName'], s['term1Total'], s['term1Paid'], s['term2Total'], s['term2Paid'], s['term3Total'], s['term3Paid']]
        for cf in s['customFees']:
            row.extend([cf['total'], cf['paid']])
        row.extend([s['totalExpected'], s['totalPaid'], s['totalPending']])
        rows.append(row)

    if format == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Status"
        ws.append(headers)
        for r in rows: ws.append(r)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename=fee_status_{studentClass}_{section}.xlsx"})
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for r in rows: writer.writerow(r)
    output.seek(0)
    return Response(content=output.getvalue(), media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=fee_status_{studentClass}_{section}.csv"})

# ==================== FEE ROUTES ====================

@api_router.get("/fees/student/{student_code}")
async def get_student_fees(student_code: str):
    student = await db.students.find_one({"studentCode": student_code}, {"_id": 0})
    if not student: raise HTTPException(status_code=404, detail="Student not found")
    payments = await db.fee_payments.find({"studentCode": student_code}, {"_id": 0}).to_list(100)
    paid_terms, paid_custom = {}, {}
    for p in payments:
        if p.get('status') in ('reverted', 'archived'): continue  # Skip reverted/archived payments
        if p.get('termNumber'):
            k = f"term{p['termNumber']}"
            paid_terms[k] = paid_terms.get(k, 0) + p['amount']
        if p.get('feeTypeId'):
            paid_custom[p['feeTypeId']] = paid_custom.get(p['feeTypeId'], 0) + p['amount']
    custom_fees = await db.fee_types.find({
        "$or": [
            {"applicableClass": student.get('studentClass', ''), "applicableSection": student.get('section', '')},
            {"applicableClass": student.get('studentClass', ''), "applicableSection": {"$in": [None, ""]}},
            {"applicableClass": {"$in": [None, ""]}, "applicableSection": {"$in": [None, ""]}},
        ]
    }, {"_id": 0}).to_list(500)
    return {"student": student, "payments": payments, "paidTerms": paid_terms, "paidCustomFees": paid_custom, "customFees": custom_fees}

@api_router.post("/fees/payment")
async def create_fee_payment(payment: FeePaymentCreate):
    # Generate sequential receipt number
    last_payment = await db.fee_payments.find({}, {"_id": 0, "receiptNumber": 1}).sort("receiptNumber", -1).to_list(1)
    if last_payment:
        try:
            last_num = int(last_payment[0]['receiptNumber'])
            next_num = last_num + 1
        except (ValueError, KeyError):
            count = await db.fee_payments.count_documents({})
            next_num = count + 1
    else:
        next_num = 1
    receipt_number = str(next_num).zfill(3)

    payment_obj = FeePayment(**payment.model_dump(), receiptNumber=receipt_number)
    doc = payment_obj.model_dump()
    doc['paymentDate'] = doc['paymentDate'].isoformat()
    await db.fee_payments.insert_one(doc)
    settings_doc = await get_wa_settings()
    student = await db.students.find_one({"studentCode": payment.studentCode}, {"_id": 0})
    if student and settings_doc:
        fee_label = f"Term {payment.termNumber}" if payment.termNumber else (payment.feeName or 'Custom Fee')
        # Build public invoice URL for WhatsApp document (view endpoint, no download header)
        backend_url = os.environ.get('REACT_APP_BACKEND_URL', '')
        invoice_url = f"{backend_url}/api/fees/invoice-view/{payment_obj.id}"
        await send_fee_paid_message(student.get('mobile', ''), invoice_url, payment.amount, fee_label, payment.studentName, settings_doc)
    return payment_obj

@api_router.get("/fees/invoice/{payment_id}")
async def download_invoice(payment_id: str):
    payment = await db.fee_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment: raise HTTPException(status_code=404, detail="Payment not found")
    student = await db.students.find_one({"studentCode": payment.get('studentCode', payment.get('rollNo', ''))}, {"_id": 0})
    if not student: student = {"studentName": payment.get('studentName', ''), "rollNo": payment.get('rollNo', ''), "studentCode": payment.get('studentCode', ''), "studentClass": "", "section": "", "fatherName": "", "mobile": ""}
    school = await db.settings.find_one({"type": "school"}, {"_id": 0})
    buf = generate_invoice_pdf(payment, student, school)
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=invoice_{payment['receiptNumber']}.pdf"})

@api_router.get("/fees/invoice-view/{payment_id}")
async def view_invoice(payment_id: str):
    """Public PDF view endpoint (no attachment header) - for WhatsApp document link"""
    payment = await db.fee_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment: raise HTTPException(status_code=404, detail="Payment not found")
    student = await db.students.find_one({"studentCode": payment.get('studentCode', payment.get('rollNo', ''))}, {"_id": 0})
    if not student: student = {"studentName": payment.get('studentName', ''), "rollNo": payment.get('rollNo', ''), "studentCode": payment.get('studentCode', ''), "studentClass": "", "section": "", "fatherName": "", "mobile": ""}
    school = await db.settings.find_one({"type": "school"}, {"_id": 0})
    buf = generate_invoice_pdf(payment, student, school)
    return StreamingResponse(buf, media_type="application/pdf")

@api_router.get("/fees/day-sheet")
async def get_day_sheet(date: Optional[str] = None):
    if not date: date = datetime.now().strftime('%Y-%m-%d')
    start = datetime.fromisoformat(f"{date}T00:00:00")
    end = datetime.fromisoformat(f"{date}T23:59:59")
    payments = await db.fee_payments.find({"paymentDate": {"$gte": start.isoformat(), "$lte": end.isoformat()}, "status": {"$nin": ["reverted", "archived"]}}, {"_id": 0}).to_list(1000)
    total = sum(p['amount'] for p in payments)
    upi_total = sum(p['amount'] for p in payments if p['paymentMode'] == 'upi')
    cash_total = sum(p['amount'] for p in payments if p['paymentMode'] == 'cash')
    return {"date": date, "payments": payments, "total": total, "upiTotal": upi_total, "cashTotal": cash_total, "count": len(payments)}

@api_router.get("/fees/export")
async def export_fees(startDate: str, endDate: str, format: str = 'csv'):
    start = datetime.fromisoformat(f"{startDate}T00:00:00")
    end = datetime.fromisoformat(f"{endDate}T23:59:59")
    payments = await db.fee_payments.find({"paymentDate": {"$gte": start.isoformat(), "$lte": end.isoformat()}}, {"_id": 0}).to_list(10000)
    if format == 'xlsx':
        wb = Workbook()
        ws = wb.active
        ws.append(['Receipt No', 'Roll No', 'Student Name', 'Fee Type', 'Amount', 'Mode', 'Date'])
        for p in payments:
            d = p['paymentDate'][:10] if isinstance(p['paymentDate'], str) else p['paymentDate'].strftime('%Y-%m-%d')
            fl = f"Term {p['termNumber']}" if p.get('termNumber') else (p.get('feeName') or 'Custom')
            ws.append([p['receiptNumber'], p['rollNo'], p['studentName'], fl, p['amount'], p['paymentMode'], d])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=fees_report.xlsx"})
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Receipt No', 'Roll No', 'Student Name', 'Fee Type', 'Amount', 'Mode', 'Date'])
    for p in payments:
        d = p['paymentDate'][:10] if isinstance(p['paymentDate'], str) else p['paymentDate'].strftime('%Y-%m-%d')
        fl = f"Term {p['termNumber']}" if p.get('termNumber') else (p.get('feeName') or 'Custom')
        writer.writerow([p['receiptNumber'], p['rollNo'], p['studentName'], fl, p['amount'], p['paymentMode'], d])
    output.seek(0)
    return Response(content=output.getvalue(), media_type="text/csv", headers={"Content-Disposition": "attachment; filename=fees_report.csv"})

@api_router.post("/fees/send-reminders")
async def send_fee_reminders():
    settings_doc = await get_wa_settings()
    upcoming = (datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')
    fee_types = await db.fee_types.find({"dueDate": {"$ne": None, "$lte": upcoming}}, {"_id": 0}).to_list(500)
    sent_count = 0
    for ft in fee_types:
        query = {}
        if ft.get('applicableClass') and ft['applicableClass']: query['studentClass'] = ft['applicableClass']
        if ft.get('applicableSection') and ft['applicableSection']: query['section'] = ft['applicableSection']
        students = await db.students.find(query, {"_id": 0}).to_list(10000)
        for student in students:
            paid = await db.fee_payments.find_one({"studentId": student['id'], "feeTypeId": ft['id']}, {"_id": 0})
            if paid: continue
            message = f"Fee Reminder: {ft['feeName']} of Rs.{ft['amount']} is due on {ft['dueDate']} for {student['studentName']}."
            result = await send_whatsapp_message(student.get('mobile', ''), message, settings_doc)
            if result.get('success'): sent_count += 1
    return {"message": f"Reminders sent to {sent_count} parents", "feeTypesChecked": len(fee_types)}

# ==================== FEE REVERT ====================

@api_router.post("/fees/revert/{payment_id}")
async def revert_fee_payment(payment_id: str):
    payment = await db.fee_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment: raise HTTPException(status_code=404, detail="Payment not found")
    if payment.get('status') == 'reverted': raise HTTPException(status_code=400, detail="Payment already reverted")
    await db.fee_payments.update_one({"id": payment_id}, {"$set": {"status": "reverted"}})
    return {"message": "Payment reverted successfully"}

# ==================== CONCESSION ROUTES ====================

@api_router.post("/concessions")
async def create_concession(data: ConcessionCreate):
    student = await db.students.find_one({"studentCode": data.studentCode}, {"_id": 0})
    if not student: raise HTTPException(status_code=404, detail="Student not found")
    obj = Concession(**data.model_dump(), studentId=student['id'], studentName=student['studentName'])
    doc = obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.concessions.insert_one(doc)
    return obj

@api_router.post("/concessions/bulk")
async def create_bulk_concession(data: BulkConcessionCreate):
    if not data.studentCodes:
        raise HTTPException(status_code=400, detail="No students selected")
    created = []
    errors = []
    for code in data.studentCodes:
        student = await db.students.find_one({"studentCode": code}, {"_id": 0})
        if not student:
            errors.append(f"{code}: not found")
            continue
        obj = Concession(
            studentId=student['id'], studentCode=code, studentName=student['studentName'],
            termNumber=data.termNumber, feeTypeId=data.feeTypeId, feeName=data.feeName,
            concessionAmount=data.concessionAmount, letterUrl=data.letterUrl, requestedBy=data.requestedBy
        )
        doc = obj.model_dump()
        doc['createdAt'] = doc['createdAt'].isoformat()
        await db.concessions.insert_one(doc)
        created.append(code)
    return {"created": len(created), "students": created, "errors": errors}

@api_router.get("/concessions")
async def get_concessions(status: Optional[str] = None):
    query = {}
    if status: query['status'] = status
    return await db.concessions.find(query, {"_id": 0}).to_list(1000)

@api_router.post("/concessions/{concession_id}/approve")
async def approve_concession(concession_id: str):
    con = await db.concessions.find_one({"id": concession_id}, {"_id": 0})
    if not con: raise HTTPException(status_code=404, detail="Concession not found")
    if con['status'] != 'pending': raise HTTPException(status_code=400, detail="Already processed")
    # Apply concession: reduce fee amount
    student = await db.students.find_one({"id": con['studentId']}, {"_id": 0})
    if not student: raise HTTPException(status_code=404, detail="Student not found")
    if con.get('termNumber'):
        field = f"feeTerm{con['termNumber']}"
        new_val = max(0, student.get(field, 0) - con['concessionAmount'])
        await db.students.update_one({"id": con['studentId']}, {"$set": {field: new_val}})
    elif con.get('feeTypeId'):
        # Store student-specific concession as a fee payment with concession flag
        await db.fee_payments.insert_one({
            "id": str(uuid.uuid4()), "studentId": con['studentId'], "studentCode": con.get('studentCode', ''),
            "rollNo": student.get('rollNo', ''), "studentName": student.get('studentName', ''),
            "feeTypeId": con['feeTypeId'], "feeName": con.get('feeName', ''),
            "amount": con['concessionAmount'], "paymentMode": "concession",
            "receiptNumber": f"CON-{str(uuid.uuid4())[:6]}", "status": "concession",
            "paymentDate": datetime.now(timezone.utc).isoformat(), "collectedBy": "System"
        })
    await db.concessions.update_one({"id": concession_id}, {"$set": {"status": "approved"}})
    return {"message": "Concession approved and applied"}

@api_router.post("/concessions/{concession_id}/reject")
async def reject_concession(concession_id: str):
    con = await db.concessions.find_one({"id": concession_id}, {"_id": 0})
    if not con: raise HTTPException(status_code=404, detail="Concession not found")
    if con['status'] != 'pending': raise HTTPException(status_code=400, detail="Already processed")
    await db.concessions.update_one({"id": concession_id}, {"$set": {"status": "rejected"}})
    return {"message": "Concession rejected"}

# ==================== LEAVE REQUEST ROUTES ====================

@api_router.post("/leave-requests")
async def create_leave_request(data: LeaveRequestCreate):
    student = await db.students.find_one({"id": data.studentId}, {"_id": 0})
    if not student:
        student = await db.students.find_one({"studentCode": data.studentCode}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    obj = LeaveRequest(
        studentId=student['id'], studentCode=student.get('studentCode', data.studentCode),
        studentName=student['studentName'], fromDate=data.fromDate, toDate=data.toDate,
        reason=data.reason, attachmentUrl=data.attachmentUrl
    )
    doc = obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.leave_requests.insert_one(doc)
    return obj

@api_router.get("/leave-requests")
async def get_leave_requests(status: Optional[str] = None, studentId: Optional[str] = None, studentClass: Optional[str] = None, section: Optional[str] = None):
    query = {}
    if status: query['status'] = status
    if studentId: query['studentId'] = studentId
    # Filter by class/section via student lookup
    if studentClass or section:
        student_query = {}
        if studentClass: student_query['studentClass'] = studentClass
        if section: student_query['section'] = section
        students = await db.students.find(student_query, {"_id": 0, "id": 1}).to_list(10000)
        student_ids = [s['id'] for s in students]
        query['studentId'] = {"$in": student_ids}
    requests = await db.leave_requests.find(query, {"_id": 0}).sort("createdAt", -1).to_list(1000)
    return requests

@api_router.post("/leave-requests/{request_id}/approve")
async def approve_leave_request(request_id: str, data: Optional[Dict] = None):
    req = await db.leave_requests.find_one({"id": request_id}, {"_id": 0})
    if not req: raise HTTPException(status_code=404, detail="Leave request not found")
    if req['status'] != 'pending': raise HTTPException(status_code=400, detail="Already processed")
    approved_by = (data or {}).get('approvedBy', 'Admin') if data else 'Admin'
    await db.leave_requests.update_one({"id": request_id}, {"$set": {"status": "approved", "approvedBy": approved_by}})
    return {"message": "Leave approved"}

@api_router.post("/leave-requests/{request_id}/reject")
async def reject_leave_request(request_id: str, data: Optional[Dict] = None):
    req = await db.leave_requests.find_one({"id": request_id}, {"_id": 0})
    if not req: raise HTTPException(status_code=404, detail="Leave request not found")
    if req['status'] != 'pending': raise HTTPException(status_code=400, detail="Already processed")
    rejected_by = (data or {}).get('rejectedBy', 'Admin') if data else 'Admin'
    await db.leave_requests.update_one({"id": request_id}, {"$set": {"status": "rejected", "rejectedBy": rejected_by}})
    return {"message": "Leave rejected"}

# ==================== MARKS ROUTES ====================

@api_router.get("/marks/sample-csv")
async def marks_sample_csv(studentClass: str, section: str):
    students = await db.students.find({"studentClass": studentClass, "section": section}, {"_id": 0}).to_list(10000)
    if not students:
        raise HTTPException(status_code=404, detail="No students found for this class & section")
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Student ID", "Name", "Exam Name", "Subject", "Marks", "Max Marks"])
    # Pre-fill student id + name; leave exam/subject/marks blank for teacher to fill
    students.sort(key=lambda s: str(s.get('rollNo', '')))
    for s in students:
        writer.writerow([s.get('studentCode', s.get('rollNo', '')), s.get('studentName', ''), '', '', '', '100'])
    output.seek(0)
    return StreamingResponse(io.BytesIO(output.getvalue().encode('utf-8')), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename=marks_template_{studentClass}_{section}.csv"})

@api_router.post("/marks/bulk")
async def create_marks_bulk(data: MarksBulkCreate):
    if not data.rows:
        raise HTTPException(status_code=400, detail="No rows provided")
    created = 0
    errors = []
    for row in data.rows:
        student = await db.students.find_one({"studentCode": row.studentCode}, {"_id": 0})
        if not student:
            errors.append(f"{row.studentCode}: not found")
            continue
        exam_name = (row.examName or data.examName).strip()
        subject = (row.subject or data.subject).strip()
        if not exam_name or not subject:
            errors.append(f"{row.studentCode}: missing exam name or subject")
            continue
        max_marks = row.maxMarks or data.maxMarks or 100
        # Upsert: replace existing record for the same student+exam+subject
        await db.marks.delete_many({
            "studentId": student['id'], "examName": exam_name, "subject": subject
        })
        entry = MarkEntry(
            studentId=student['id'], studentCode=row.studentCode,
            studentName=student.get('studentName', ''),
            studentClass=data.studentClass, section=data.section,
            examName=exam_name, subject=subject,
            marks=float(row.marks), maxMarks=float(max_marks),
            recordedBy=data.recordedBy or 'Teacher'
        )
        doc = entry.model_dump()
        doc['recordedOn'] = doc['recordedOn'].isoformat()
        await db.marks.insert_one(doc)
        created += 1
    return {"created": created, "errors": errors}

@api_router.get("/marks")
async def get_marks(studentId: Optional[str] = None, studentClass: Optional[str] = None,
                    section: Optional[str] = None, examName: Optional[str] = None,
                    subject: Optional[str] = None):
    query = {}
    if studentId: query['studentId'] = studentId
    if studentClass: query['studentClass'] = studentClass
    if section: query['section'] = section
    if examName: query['examName'] = examName
    if subject: query['subject'] = subject
    return await db.marks.find(query, {"_id": 0}).sort("recordedOn", -1).to_list(10000)

@api_router.get("/marks/distinct")
async def get_marks_distinct():
    """Return distinct exams and subjects available."""
    exams = await db.marks.distinct("examName")
    subjects = await db.marks.distinct("subject")
    classes_in_marks = await db.marks.distinct("studentClass")
    return {"exams": sorted([e for e in exams if e]), "subjects": sorted([s for s in subjects if s]),
            "classes": sorted([c for c in classes_in_marks if c])}

@api_router.get("/marks/stats")
async def get_marks_stats(studentClass: Optional[str] = None, section: Optional[str] = None,
                          examName: Optional[str] = None, subject: Optional[str] = None,
                          compareExamA: Optional[str] = None, compareExamB: Optional[str] = None):
    """Aggregated statistics for analytics dashboards."""
    base_query = {}
    if studentClass: base_query['studentClass'] = studentClass
    if section: base_query['section'] = section
    if subject: base_query['subject'] = subject

    # For single exam filter
    single_q = dict(base_query)
    if examName: single_q['examName'] = examName

    marks_records = await db.marks.find(single_q, {"_id": 0}).to_list(10000)

    # Overall metrics
    pct_list = [(m['marks'] / m['maxMarks'] * 100) if m.get('maxMarks') else 0 for m in marks_records]
    overall = {
        "totalEntries": len(marks_records),
        "averagePct": round(sum(pct_list) / len(pct_list), 2) if pct_list else 0,
        "highestPct": round(max(pct_list), 2) if pct_list else 0,
        "lowestPct": round(min(pct_list), 2) if pct_list else 0,
        "passCount": sum(1 for p in pct_list if p >= 33),
        "failCount": sum(1 for p in pct_list if p < 33),
    }

    # Group by subject -> average pct
    subj_map = {}
    for m in marks_records:
        s = m.get('subject', 'Unknown')
        if s not in subj_map: subj_map[s] = []
        pct = (m['marks'] / m['maxMarks'] * 100) if m.get('maxMarks') else 0
        subj_map[s].append(pct)
    bySubject = [{"subject": k, "average": round(sum(v) / len(v), 2), "count": len(v)} for k, v in subj_map.items()]
    bySubject.sort(key=lambda x: x['subject'])

    # Grade distribution buckets
    buckets = {"A+ (90+)": 0, "A (75-89)": 0, "B (60-74)": 0, "C (45-59)": 0, "D (33-44)": 0, "F (<33)": 0}
    for p in pct_list:
        if p >= 90: buckets["A+ (90+)"] += 1
        elif p >= 75: buckets["A (75-89)"] += 1
        elif p >= 60: buckets["B (60-74)"] += 1
        elif p >= 45: buckets["C (45-59)"] += 1
        elif p >= 33: buckets["D (33-44)"] += 1
        else: buckets["F (<33)"] += 1
    grades = [{"grade": k, "count": v} for k, v in buckets.items()]

    # Top students (top 10)
    student_map = {}
    for m in marks_records:
        sid = m['studentId']
        if sid not in student_map:
            student_map[sid] = {"studentId": sid, "studentName": m.get('studentName', ''),
                                "studentCode": m.get('studentCode', ''), "marks": 0, "max": 0}
        student_map[sid]['marks'] += m['marks']
        student_map[sid]['max'] += m.get('maxMarks', 100)
    students_stats = []
    for v in student_map.values():
        pct = (v['marks'] / v['max'] * 100) if v['max'] else 0
        students_stats.append({**v, "pct": round(pct, 2)})
    students_stats.sort(key=lambda x: x['pct'], reverse=True)
    topStudents = students_stats[:10]

    # Compare two exams
    compare = None
    if compareExamA and compareExamB:
        a_q = dict(base_query); a_q['examName'] = compareExamA
        b_q = dict(base_query); b_q['examName'] = compareExamB
        rec_a = await db.marks.find(a_q, {"_id": 0}).to_list(10000)
        rec_b = await db.marks.find(b_q, {"_id": 0}).to_list(10000)
        def avg_by_subject(records):
            mp = {}
            for r in records:
                s = r.get('subject', 'Unknown')
                if s not in mp: mp[s] = []
                pct = (r['marks'] / r['maxMarks'] * 100) if r.get('maxMarks') else 0
                mp[s].append(pct)
            return {k: round(sum(v) / len(v), 2) for k, v in mp.items()}
        a_subj = avg_by_subject(rec_a)
        b_subj = avg_by_subject(rec_b)
        all_subj = sorted(set(list(a_subj.keys()) + list(b_subj.keys())))
        compare = [{"subject": s, compareExamA: a_subj.get(s, 0), compareExamB: b_subj.get(s, 0)} for s in all_subj]

    return {"overall": overall, "bySubject": bySubject, "grades": grades,
            "topStudents": topStudents, "compare": compare}


# ==================== EXPENSE ROUTES ====================

@api_router.post("/expenses", response_model=Expense)
async def create_expense(expense: ExpenseCreate):
    obj = Expense(**expense.model_dump())
    doc = obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.expenses.insert_one(doc)
    return obj

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(startDate: Optional[str] = None, endDate: Optional[str] = None):
    query = {}
    if startDate and endDate: query['date'] = {'$gte': startDate, '$lte': endDate}
    expenses = await db.expenses.find(query, {"_id": 0}).to_list(1000)
    for e in expenses:
        if isinstance(e.get('createdAt'), str): e['createdAt'] = datetime.fromisoformat(e['createdAt'])
    return expenses

# ==================== SETTINGS ROUTES ====================

@api_router.get("/settings/whatsapp")
async def get_whatsapp_settings():
    settings = await db.settings.find_one({"type": "whatsapp"}, {"_id": 0})
    if not settings: return {"phoneNumberId": "", "accessToken": ""}
    return settings

@api_router.put("/settings/whatsapp")
async def update_whatsapp_settings(settings: WhatsAppSettings):
    await db.settings.update_one({"type": "whatsapp"}, {"$set": {"phoneNumberId": settings.phoneNumberId, "accessToken": settings.accessToken}}, upsert=True)
    return {"message": "Settings updated"}

@api_router.get("/settings/school")
async def get_school_settings():
    settings = await db.settings.find_one({"type": "school"}, {"_id": 0})
    if not settings: return {"schoolName": "SchoolPro", "schoolAddress": "", "logoUrl": ""}
    return settings

@api_router.put("/settings/school")
async def update_school_settings(data: SchoolSettings):
    await db.settings.update_one({"type": "school"}, {"$set": {"schoolName": data.schoolName, "schoolAddress": data.schoolAddress, "logoUrl": data.logoUrl or ""}}, upsert=True)
    return {"message": "School settings updated"}

# ==================== FILE UPLOAD ====================

@api_router.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    content = await file.read()
    base64_content = base64.b64encode(content).decode('utf-8')
    return {"url": f"data:{file.content_type};base64,{base64_content}", "filename": file.filename}

# ==================== INVENTORY ROUTES ====================

@api_router.post("/inventory")
async def create_inventory_item(item: InventoryItemCreate):
    obj = InventoryItem(**item.model_dump())
    doc = obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.inventory.insert_one(doc)
    return obj

@api_router.get("/inventory")
async def get_inventory(category: Optional[str] = None):
    query = {}
    if category: query['category'] = category
    return await db.inventory.find(query, {"_id": 0}).to_list(1000)

@api_router.put("/inventory/{item_id}")
async def update_inventory_item(item_id: str, data: InventoryItemCreate):
    result = await db.inventory.update_one({"id": item_id}, {"$set": data.model_dump()})
    if result.matched_count == 0: raise HTTPException(status_code=404, detail="Item not found")
    return await db.inventory.find_one({"id": item_id}, {"_id": 0})

@api_router.delete("/inventory/{item_id}")
async def delete_inventory_item(item_id: str):
    result = await db.inventory.delete_one({"id": item_id})
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="Item not found")
    return {"message": "Item deleted"}

@api_router.post("/inventory/issue")
async def issue_inventory(data: InventoryIssueCreate):
    item = await db.inventory.find_one({"id": data.itemId}, {"_id": 0})
    if not item: raise HTTPException(status_code=404, detail="Item not found")
    if item['quantity'] < data.quantity: raise HTTPException(status_code=400, detail=f"Insufficient stock. Available: {item['quantity']}")
    student = await db.students.find_one({"studentCode": data.studentCode}, {"_id": 0})
    if not student: raise HTTPException(status_code=404, detail="Student not found")
    # Deduct stock
    await db.inventory.update_one({"id": data.itemId}, {"$inc": {"quantity": -data.quantity}})
    issue = InventoryIssue(itemId=data.itemId, itemName=item['itemName'], studentId=student['id'],
                           studentCode=data.studentCode, rollNo=student.get('rollNo', ''), studentName=student['studentName'], quantity=data.quantity, date=data.date)
    doc = issue.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.inventory_issues.insert_one(doc)
    return issue

@api_router.get("/inventory/issues")
async def get_inventory_issues(studentId: Optional[str] = None):
    query = {}
    if studentId: query['studentId'] = studentId
    return await db.inventory_issues.find(query, {"_id": 0}).to_list(1000)

# ==================== EVENT ROUTES ====================

@api_router.post("/events")
async def create_event(event: EventCreate):
    obj = Event(**event.model_dump())
    doc = obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.events.insert_one(doc)
    # Send WhatsApp notification to all students if enabled
    if event.sendNotification:
        settings_doc = await get_wa_settings()
        if settings_doc:
            students = await db.students.find({}, {"_id": 0, "mobile": 1}).to_list(10000)
            for student in students:
                if student.get('mobile'):
                    await send_event_message(student['mobile'], event.title, event.date, settings_doc)
    return obj

@api_router.get("/events")
async def get_events(month: Optional[str] = None):
    query = {}
    if month: query['date'] = {'$regex': f'^{month}'}
    return await db.events.find(query, {"_id": 0}).to_list(1000)

@api_router.put("/events/{event_id}")
async def update_event(event_id: str, data: EventCreate):
    result = await db.events.update_one({"id": event_id}, {"$set": data.model_dump()})
    if result.matched_count == 0: raise HTTPException(status_code=404, detail="Event not found")
    return await db.events.find_one({"id": event_id}, {"_id": 0})

@api_router.delete("/events/{event_id}")
async def delete_event(event_id: str):
    result = await db.events.delete_one({"id": event_id})
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="Event not found")
    return {"message": "Event deleted"}

# ==================== HOMEWORK ROUTES ====================

@api_router.post("/homework")
async def create_homework(hw: HomeworkCreate):
    obj = Homework(**hw.model_dump())
    doc = obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.homework.insert_one(doc)
    return obj

@api_router.get("/homework")
async def get_homework(studentClass: Optional[str] = None, section: Optional[str] = None):
    query = {}
    if studentClass: query['studentClass'] = studentClass
    if section: query['section'] = section
    return await db.homework.find(query, {"_id": 0}).to_list(1000)

@api_router.put("/homework/{hw_id}")
async def update_homework(hw_id: str, data: HomeworkCreate):
    result = await db.homework.update_one({"id": hw_id}, {"$set": data.model_dump()})
    if result.matched_count == 0: raise HTTPException(status_code=404, detail="Homework not found")
    return await db.homework.find_one({"id": hw_id}, {"_id": 0})

@api_router.delete("/homework/{hw_id}")
async def delete_homework(hw_id: str):
    result = await db.homework.delete_one({"id": hw_id})
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="Homework not found")
    return {"message": "Homework deleted"}

# ==================== STAFF ROUTES ====================

@api_router.post("/staff")
async def create_staff(data: StaffCreate):
    existing = await db.staff.find_one({"username": data.username}, {"_id": 0})
    if existing: raise HTTPException(status_code=400, detail="Username already exists")
    obj = Staff(**data.model_dump())
    doc = obj.model_dump()
    doc['createdAt'] = doc['createdAt'].isoformat()
    await db.staff.insert_one(doc)
    return {k: v for k, v in obj.model_dump().items() if k != 'password'}

@api_router.get("/staff")
async def get_staff():
    staff = await db.staff.find({}, {"_id": 0}).to_list(500)
    return [{k: v for k, v in s.items() if k != 'password'} for s in staff]

@api_router.put("/staff/{staff_id}")
async def update_staff(staff_id: str, data: StaffUpdate):
    update_dict = {k: v for k, v in data.model_dump().items() if v is not None}
    if not update_dict: return await db.staff.find_one({"id": staff_id}, {"_id": 0, "password": 0})
    result = await db.staff.update_one({"id": staff_id}, {"$set": update_dict})
    if result.matched_count == 0: raise HTTPException(status_code=404, detail="Staff not found")
    updated = await db.staff.find_one({"id": staff_id}, {"_id": 0})
    return {k: v for k, v in updated.items() if k != 'password'}

@api_router.delete("/staff/{staff_id}")
async def delete_staff(staff_id: str):
    result = await db.staff.delete_one({"id": staff_id})
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="Staff not found")
    return {"message": "Staff deleted"}

# ==================== STUDENT DETAIL ====================

@api_router.get("/students/{student_id}/detail")
async def get_student_detail(student_id: str):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student: raise HTTPException(status_code=404, detail="Student not found")
    attendance = await db.attendance.find({"studentId": student_id}, {"_id": 0}).to_list(10000)
    total_days = len(attendance)
    present_days = sum(1 for a in attendance if a['status'] == 'present')
    absent_days = sum(1 for a in attendance if a['status'] == 'absent')
    payments = await db.fee_payments.find({"studentId": student_id}, {"_id": 0}).to_list(100)
    custom_fees = await db.fee_types.find({
        "$or": [
            {"applicableClass": student.get('studentClass', ''), "applicableSection": student.get('section', '')},
            {"applicableClass": student.get('studentClass', ''), "applicableSection": {"$in": [None, ""]}},
            {"applicableClass": {"$in": [None, ""]}, "applicableSection": {"$in": [None, ""]}},
        ]
    }, {"_id": 0}).to_list(500)
    paid_terms, paid_custom = {}, {}
    for p in payments:
        if p.get('status') in ('reverted', 'archived'): continue
        if p.get('termNumber'): k = f"term{p['termNumber']}"; paid_terms[k] = paid_terms.get(k, 0) + p['amount']
        if p.get('feeTypeId'): paid_custom[p['feeTypeId']] = paid_custom.get(p['feeTypeId'], 0) + p['amount']
    # Inventory issued
    inventory_issued = await db.inventory_issues.find({"studentId": student_id}, {"_id": 0}).to_list(500)
    return {
        "student": student, "attendance": attendance,
        "attendanceStats": {"totalDays": total_days, "presentDays": present_days, "absentDays": absent_days,
                            "percentage": round(present_days / total_days * 100, 1) if total_days > 0 else 0},
        "payments": payments, "paidTerms": paid_terms, "paidCustomFees": paid_custom, "customFees": custom_fees,
        "inventoryIssued": inventory_issued,
        "promotionHistory": student.get('promotionHistory', []),
    }

# ==================== PARENT PORTAL ====================

@api_router.get("/parent/dashboard/{student_id}")
async def parent_dashboard(student_id: str):
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student: raise HTTPException(status_code=404, detail="Student not found")
    attendance = await db.attendance.find({"studentId": student_id}, {"_id": 0}).to_list(10000)
    total_days = len(attendance)
    present_days = sum(1 for a in attendance if a['status'] == 'present')
    absent_days = sum(1 for a in attendance if a['status'] == 'absent')
    payments = await db.fee_payments.find({"studentId": student_id}, {"_id": 0}).to_list(100)
    events = await db.events.find({}, {"_id": 0}).to_list(100)
    homework_cutoff = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    homework = await db.homework.find({"studentClass": student.get('studentClass', ''), "section": student.get('section', ''), "createdAt": {"$gte": homework_cutoff}}, {"_id": 0}).to_list(100)
    # Fallback: if createdAt is stored as ISO string, also get by dueDate
    if not homework:
        homework = await db.homework.find({"studentClass": student.get('studentClass', ''), "section": student.get('section', ''), "dueDate": {"$gte": homework_cutoff}}, {"_id": 0}).to_list(100)
    # Fee structure
    custom_fees = await db.fee_types.find({
        "$or": [
            {"applicableClass": student.get('studentClass', ''), "applicableSection": student.get('section', '')},
            {"applicableClass": student.get('studentClass', ''), "applicableSection": {"$in": [None, ""]}},
            {"applicableClass": {"$in": [None, ""]}, "applicableSection": {"$in": [None, ""]}},
        ]
    }, {"_id": 0}).to_list(500)
    paid_terms, paid_custom = {}, {}
    for p in payments:
        if p.get('status') in ('reverted', 'archived'): continue
        if p.get('termNumber'):
            k = f"term{p['termNumber']}"
            paid_terms[k] = paid_terms.get(k, 0) + p['amount']
        if p.get('feeTypeId'):
            paid_custom[p['feeTypeId']] = paid_custom.get(p['feeTypeId'], 0) + p['amount']
    return {
        "student": {k: v for k, v in student.items() if k != 'parentPassword'},
        "attendanceStats": {"totalDays": total_days, "presentDays": present_days, "absentDays": absent_days,
                            "percentage": round(present_days / total_days * 100, 1) if total_days > 0 else 0},
        "recentAttendance": attendance[-30:] if attendance else [],
        "fullAttendance": attendance,
        "payments": payments, "events": events, "homework": homework,
        "marks": await db.marks.find({"studentId": student['id']}, {"_id": 0}).sort("recordedOn", -1).to_list(1000),
        "feeStructure": {
            "term1": {"total": student.get('feeTerm1', 0), "paid": paid_terms.get('term1', 0)},
            "term2": {"total": student.get('feeTerm2', 0), "paid": paid_terms.get('term2', 0)},
            "term3": {"total": student.get('feeTerm3', 0), "paid": paid_terms.get('term3', 0)},
            "customFees": [{"id": cf['id'], "feeName": cf['feeName'], "total": cf['amount'], "paid": paid_custom.get(cf['id'], 0), "dueDate": cf.get('dueDate')} for cf in custom_fees],
        },
        "paidTerms": paid_terms, "paidCustomFees": paid_custom, "customFees": custom_fees,
    }

# ==================== DASHBOARD STATS ====================

@api_router.get("/stats/dashboard")
async def get_dashboard_stats():
    total_students = await db.students.count_documents({})
    today = datetime.now().strftime('%Y-%m-%d')
    today_present = await db.attendance.count_documents({"date": today, "status": "present"})
    today_absent = await db.attendance.count_documents({"date": today, "status": "absent"})
    pipeline = [{"$group": {"_id": None, "total": {"$sum": "$amount"}}}]
    fees_result = await db.fee_payments.aggregate(pipeline).to_list(1)
    total_fees = fees_result[0]['total'] if fees_result else 0
    students = await db.students.find({}, {"_id": 0, "feeTerm1": 1, "feeTerm2": 1, "feeTerm3": 1}).to_list(10000)
    total_expected = sum(s.get('feeTerm1', 0) + s.get('feeTerm2', 0) + s.get('feeTerm3', 0) for s in students)
    return {"totalStudents": total_students, "presentToday": today_present, "absentToday": today_absent,
            "totalFeesCollected": total_fees, "pendingFees": total_expected - total_fees}

@api_router.get("/")
async def root():
    return {"message": "SchoolPro API"}

app.include_router(api_router)

app.add_middleware(CORSMiddleware, allow_credentials=True,
                   allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
                   allow_methods=["*"], allow_headers=["*"])

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
